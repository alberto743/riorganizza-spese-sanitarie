# SPDX-FileCopyrightText: 2026 Alberto P
# SPDX-License-Identifier: MPL-2.0
"""Scrittura .xlsx (openpyxl)."""

from __future__ import annotations

from datetime import date
from pathlib import Path

from .csv_sorgente import Pagamento
from .formule import COLORE_INTESTAZIONE, FORMATO_DATA, FORMATO_VALUTA, INTESTAZIONI, _righe_output


def scrivi_xlsx(pagamenti: list[Pagamento], percorso: Path) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    font_intestazione = Font(name="Arial", bold=True, color="FFFFFF")
    fill_intestazione = PatternFill("solid", fgColor=COLORE_INTESTAZIONE)
    font_normale = Font(name="Arial")
    font_formula = Font(name="Arial", bold=True)
    allinea_centro = Alignment(horizontal="center", vertical="center", wrap_text=True)
    bordo_sottile = Border(*(Side(style="thin", color="D9D9D9"),) * 4)

    wb = Workbook()
    ws = wb.active
    ws.title = "Spese sanitarie"

    for c, titolo in enumerate(INTESTAZIONI, start=1):
        cella = ws.cell(row=1, column=c, value=titolo)
        cella.font = font_intestazione
        cella.fill = fill_intestazione
        cella.alignment = allinea_centro
        cella.border = bordo_sottile
    ws.freeze_panes = "A2"

    r = 2
    for data_pag, emesso_da, formula_tot, _v_tot, formula_detr, _v_detr in _righe_output(pagamenti):
        cella_data = ws.cell(row=r, column=1, value=data_pag)
        cella_data.font = font_normale
        cella_data.alignment = allinea_centro
        if isinstance(data_pag, date):
            cella_data.number_format = FORMATO_DATA
        # se e' rimasta una stringa (data non riconosciuta nel CSV originale)
        # la cella resta testo, per non perdere l'informazione originale.

        ws.cell(row=r, column=2, value=emesso_da).font = font_normale

        cella_totale = ws.cell(row=r, column=3, value=formula_tot)
        cella_totale.font = font_formula
        cella_totale.number_format = FORMATO_VALUTA

        cella_detraibile = ws.cell(row=r, column=4, value=formula_detr)
        cella_detraibile.font = font_formula
        cella_detraibile.number_format = FORMATO_VALUTA

        for c in range(1, len(INTESTAZIONI) + 1):
            ws.cell(row=r, column=c).border = bordo_sottile
        r += 1

    ultima_riga = r - 1
    riga_totali = r
    ws.cell(row=riga_totali, column=2, value="Totale").font = Font(name="Arial", bold=True)
    for col, lettera in ((3, "C"), (4, "D")):
        cella = ws.cell(row=riga_totali, column=col, value=f"=SUM({lettera}2:{lettera}{ultima_riga})")
        cella.font = Font(name="Arial", bold=True)
        cella.number_format = FORMATO_VALUTA
        cella.border = Border(top=Side(style="thin"))

    ws.column_dimensions["A"].width = 13
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18

    percorso.parent.mkdir(parents=True, exist_ok=True)
    wb.save(percorso)
