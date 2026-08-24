# SPDX-FileCopyrightText: 2026 Alberto P
# SPDX-License-Identifier: MPL-2.0
from datetime import date

from riorganizza_spese_sanitarie.csv_sorgente import Pagamento, Sottovoce
from riorganizza_spese_sanitarie.formule import INTESTAZIONI
from riorganizza_spese_sanitarie.scrittura_ods import scrivi_ods
from riorganizza_spese_sanitarie.scrittura_xlsx import scrivi_xlsx

PAGAMENTI = [
    Pagamento(
        data_pagamento=date(2026, 2, 1),
        emesso_da="=cmd|'/c calc'!A1",  # tentativo di formula-injection
        sottovoci=[Sottovoce("Farmaco", 12.50, True)],
    ),
]


def test_scrivi_xlsx_intestazione_e_formula(tmp_path):
    from openpyxl import load_workbook

    percorso = tmp_path / "out.xlsx"
    scrivi_xlsx(PAGAMENTI, percorso)

    wb = load_workbook(percorso)
    ws = wb.active
    assert [c.value for c in ws[1]] == INTESTAZIONI

    cella_totale = ws.cell(row=2, column=3)
    assert cella_totale.value == "=12.5"
    assert cella_totale.data_type == "f"


def test_scrivi_xlsx_emesso_da_non_diventa_formula(tmp_path):
    from openpyxl import load_workbook

    percorso = tmp_path / "out.xlsx"
    scrivi_xlsx(PAGAMENTI, percorso)

    wb = load_workbook(percorso)
    ws = wb.active
    cella_emesso = ws.cell(row=2, column=2)

    assert cella_emesso.data_type == "s"
    assert cella_emesso.value == "=cmd|'/c calc'!A1"


def test_scrivi_ods_intestazione_e_formula(tmp_path):
    from odf.opendocument import load
    from odf.table import TableCell, TableRow
    from odf.text import P

    percorso = tmp_path / "out.ods"
    scrivi_ods(PAGAMENTI, percorso)

    doc = load(str(percorso))
    tabella = doc.spreadsheet.getElementsByType(__import__("odf.table", fromlist=["Table"]).Table)[0]
    righe = tabella.getElementsByType(TableRow)

    celle_intestazione = righe[0].getElementsByType(TableCell)
    testi_intestazione = [str(c.getElementsByType(P)[0]) for c in celle_intestazione]
    assert testi_intestazione == INTESTAZIONI

    cella_totale = righe[1].getElementsByType(TableCell)[2]
    assert cella_totale.getAttribute("formula") == "of:=12.5"


def test_scrivi_ods_emesso_da_resta_stringa(tmp_path):
    from odf.opendocument import load
    from odf.table import TableCell, TableRow
    from odf.text import P

    percorso = tmp_path / "out.ods"
    scrivi_ods(PAGAMENTI, percorso)

    doc = load(str(percorso))
    tabella = doc.spreadsheet.getElementsByType(__import__("odf.table", fromlist=["Table"]).Table)[0]
    righe = tabella.getElementsByType(TableRow)

    cella_emesso = righe[1].getElementsByType(TableCell)[1]
    assert cella_emesso.getAttribute("valuetype") == "string"
    assert cella_emesso.getAttribute("formula") is None
    assert str(cella_emesso.getElementsByType(P)[0]) == "=cmd|'/c calc'!A1"
